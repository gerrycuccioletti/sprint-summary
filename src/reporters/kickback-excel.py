#!/usr/bin/env python3
# kickback-excel.py — generates kickback ratio Excel report via openpyxl
import json, sys, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

def hex_fill(hex_color):
    return PatternFill('solid', start_color=hex_color, end_color=hex_color)

def thin_border():
    s = Side(style='thin', color='D0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def ratio_fill(ratio):
    if ratio == 0:    return hex_fill('D1FAE5')   # green
    if ratio <= 10:   return hex_fill('FEF9C3')   # yellow
    if ratio <= 25:   return hex_fill('FFEDD5')   # orange
    return hex_fill('FEE2E2')                      # red

def ratio_font(ratio):
    if ratio == 0:    return Font(bold=True, color='065F46')
    if ratio <= 10:   return Font(bold=True, color='854D0E')
    if ratio <= 25:   return Font(bold=True, color='9A3412')
    return Font(bold=True, color='991B1B')

LABELS = {
    'CR':   {'d1': '→ Code Review',   'd2': 'In Code Review',   'back': 'Kicked Back (→Dev)',    'status': 'Code Review',   'kickStatus': 'Dev In Progress'},
    'QA':   {'d1': '→ In QA',         'd2': 'In QA',            'back': 'Kicked Back (→ReOpen)', 'status': 'In QA',         'kickStatus': 'ReOpen'},
    'RFSO': {'d1': '→ Ready for SO',  'd2': 'In Ready for SO',  'back': 'Kicked Back (→ReOpen)', 'status': 'Ready for SO',  'kickStatus': 'ReOpen'},
}

def write_report(config):
    rt       = config['reportType']
    results  = config['results']
    fromDate = config['fromDate']
    toDate   = config['toDate']
    lbl      = LABELS.get(rt, LABELS['CR'])
    out_dir  = config.get('outputDir', './output')
    os.makedirs(out_dir, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f'{rt} Kickback Ratio Report — {fromDate} to {toDate}'
    ws['A1'].font      = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill      = hex_fill('1E3A5F')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Generated
    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font      = Font(name='Arial', size=9, color='64748B', italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Project', lbl['d1'], lbl['d2'], lbl['back'], 'Ratio 1 (%)', 'Ratio 2 (%)', 'D1 JQL', 'D2 JQL', 'Numerator JQL']
    hdr_fill = hex_fill('2563EB')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = thin_border()
    ws.row_dimensions[4].height = 30

    # Data rows
    valid   = [r for r in results if not r.get('error')]
    errored = [r for r in results if r.get('error')]
    row = 5

    for r in valid:
        d1    = r.get('denominator', 0)
        d2    = r.get('denominator2', 0)
        num   = r.get('numerator', 0)
        rat1  = r.get('ratio', 0)
        rat2  = r.get('ratio2', 0)
        fslash = fromDate.replace('-', '/')
        tslash = toDate.replace('-', '/')
        jql_d1  = f'project = "{r["projectKey"]}" AND status changed TO "{lbl["status"]}" DURING ("{fslash}", "{tslash}")'
        jql_d2  = f'project = "{r["projectKey"]}" AND status was "{lbl["status"]}" DURING ("{fslash}", "{tslash}")'
        jql_num = f'project = "{r["projectKey"]}" AND status changed FROM "{lbl["status"]}" TO "{lbl["kickStatus"]}" DURING ("{fslash}", "{tslash}")'

        row_data = [r['projectKey'], d1, d2, num, rat1, rat2, jql_d1, jql_d2, jql_num]
        alt_fill = hex_fill('F8FAFC') if row % 2 == 0 else hex_fill('FFFFFF')

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name='Arial', size=10)
            c.border    = thin_border()
            c.alignment = Alignment(vertical='center')
            if col in (5, 6):   # Ratio columns
                c.number_format = '0.0"%"'
                c.fill  = ratio_fill(val)
                c.font  = ratio_font(val)
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif col in (2, 3, 4):
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = alt_fill
            elif col in (7, 8, 9):
                c.font = Font(name='Courier New', size=8, color='334155')
                c.alignment = Alignment(wrap_text=True, vertical='center')
                c.fill = hex_fill('F1F5F9')
            else:
                c.font = Font(name='Arial', bold=True, size=10)
                c.fill = alt_fill
        ws.row_dimensions[row].height = 22
        row += 1

    # Error rows
    for r in errored:
        ws.cell(row=row, column=1, value=r['projectKey']).font = Font(name='Arial', bold=True, color='991B1B')
        c = ws.cell(row=row, column=2, value=f'ERROR: {r["error"]}')
        c.font = Font(name='Arial', color='991B1B', italic=True)
        ws.merge_cells(f'B{row}:H{row}')
        row += 1

    # Totals row
    if valid:
        tot_d1   = sum(r.get('denominator',  0) for r in valid)
        tot_d2   = sum(r.get('denominator2', 0) for r in valid)
        tot_num  = sum(r.get('numerator',    0) for r in valid)
        tot_rat1 = round((tot_num / tot_d1  * 100), 1) if tot_d1  > 0 else 0
        tot_rat2 = round((tot_num / tot_d2  * 100), 1) if tot_d2  > 0 else 0

        tot_data = ['TOTAL', tot_d1, tot_d2, tot_num, tot_rat1, tot_rat2, '', '', '']
        for col, val in enumerate(tot_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font   = Font(name='Arial', bold=True, size=10, color='1E3A5F')
            c.fill   = hex_fill('DBEAFE')
            c.border = thin_border()
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col in (5, 6):
                c.number_format = '0.0"%"'
                c.fill = ratio_fill(tot_rat1 if col == 5 else tot_rat2)
                c.font = ratio_font(tot_rat1 if col == 5 else tot_rat2)
        ws.row_dimensions[row].height = 24

    # Column widths
    widths = [14, 14, 14, 14, 12, 12, 55, 55, 65]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A5'

    # ── Sheet 2: Kickback Detail ─────────────────────────────────────────
    ws2 = wb.create_sheet('Kickback Detail')
    ws2.merge_cells('A1:F1')
    ws2['A1'] = f'{rt} Kickback Detail — {fromDate} to {toDate}'
    ws2['A1'].font      = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws2['A1'].fill      = hex_fill('1E3A5F')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 26

    detail_hdrs = ['Project', 'Issue Key', 'Summary', 'Assignee']
    for col, h in enumerate(detail_hdrs, 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill      = hex_fill('2563EB')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = thin_border()
    ws2.row_dimensions[3].height = 22

    dr = 4
    for r in valid:
        for kb in r.get('details', {}).get('kickedBack', []):
            row_vals = [
                r['projectKey'],
                kb.get('key', ''),
                kb.get('summary', ''),
                kb.get('author', ''),
            ]
            alt = hex_fill('F8FAFC') if dr % 2 == 0 else hex_fill('FFFFFF')
            for col, val in enumerate(row_vals, 1):
                c = ws2.cell(row=dr, column=col, value=val)
                c.font   = Font(name='Arial', size=10)
                c.fill   = alt
                c.border = thin_border()
                c.alignment = Alignment(vertical='center', wrap_text=(col == 3))
                if col == 2:
                    c.font = Font(name='Courier New', size=10, bold=True, color='2563EB')
            ws2.row_dimensions[dr].height = 20
            dr += 1

    if dr == 4:
        ws2.cell(row=4, column=1, value='No kickbacks found in this period').font = Font(name='Arial', italic=True, color='94A3B8')

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 55
    ws2.column_dimensions['D'].width = 22
    ws2.freeze_panes = 'A4'

    # ── Sheet 3: D1 Issues ───────────────────────────────────────────────
    ws3 = wb.create_sheet(f'D1 — {lbl["d1"]}')
    ws3.merge_cells('A1:C1')
    ws3['A1'] = f'D1: {lbl["d1"]} — {fromDate} to {toDate}'
    ws3['A1'].font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws3['A1'].fill = hex_fill('1E3A5F')
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 26

    for col, h in enumerate(['Project', 'Issue Key', 'Summary'], 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = hex_fill('2563EB')
        c.border = thin_border()

    d1r = 4
    d1_key = 'toCodeReview' if rt == 'CR' else ('toInQA' if rt == 'QA' else 'toRFSO')
    for r in valid:
        for issue in r.get('details', {}).get(d1_key, []):
            alt = hex_fill('F8FAFC') if d1r % 2 == 0 else hex_fill('FFFFFF')
            for col, val in enumerate([r['projectKey'], issue.get('key',''), issue.get('summary','')], 1):
                c = ws3.cell(row=d1r, column=col, value=val)
                c.font = Font(name='Arial' if col != 2 else 'Courier New', size=10,
                              bold=(col==2), color=('2563EB' if col==2 else '000000'))
                c.fill = alt
                c.border = thin_border()
                c.alignment = Alignment(vertical='center', wrap_text=(col==3))
            ws3.row_dimensions[d1r].height = 20
            d1r += 1

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 60
    ws3.freeze_panes = 'A4'

    # ── Sheet 4: D2 Issues ───────────────────────────────────────────────
    d2_key   = 'inCodeReview' if rt == 'CR' else ('inQA' if rt == 'QA' else 'inRFSO')
    d2_label = lbl['d2']
    ws4 = wb.create_sheet(f'D2 — {d2_label}')
    ws4.merge_cells('A1:C1')
    ws4['A1'] = f'D2: {d2_label} — {fromDate} to {toDate}'
    ws4['A1'].font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws4['A1'].fill = hex_fill('1E3A5F')
    ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 26

    for col, h in enumerate(['Project', 'Issue Key', 'Summary'], 1):
        c = ws4.cell(row=3, column=col, value=h)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = hex_fill('2563EB')
        c.border = thin_border()

    d2r = 4
    for r in valid:
        for issue in r.get('details', {}).get(d2_key, []):
            alt = hex_fill('F8FAFC') if d2r % 2 == 0 else hex_fill('FFFFFF')
            for col, val in enumerate([r['projectKey'], issue.get('key',''), issue.get('summary','')], 1):
                c = ws4.cell(row=d2r, column=col, value=val)
                c.font = Font(name='Arial' if col != 2 else 'Courier New', size=10,
                              bold=(col==2), color=('2563EB' if col==2 else '000000'))
                c.fill = alt
                c.border = thin_border()
                c.alignment = Alignment(vertical='center', wrap_text=(col==3))
            ws4.row_dimensions[d2r].height = 20
            d2r += 1

    if d2r == 4:
        ws4.cell(row=4, column=1, value='No issues found in this period').font = Font(name='Arial', italic=True, color='94A3B8')

    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 14
    ws4.column_dimensions['C'].width = 60
    ws4.freeze_panes = 'A4'

    # Save
    filename = f'{rt.lower()}-kickback-{fromDate}-{toDate}.xlsx'
    filepath = os.path.join(out_dir, filename)
    wb.save(filepath)
    return filepath

if __name__ == '__main__':
    try:
        cfg_path = sys.argv[1]
        with open(cfg_path, encoding='utf-8') as f:
            config = json.load(f)
        filepath = write_report(config)
        print(json.dumps({'filePath': filepath}))
    except Exception as e:
        print(json.dumps({'error': str(e)}))
